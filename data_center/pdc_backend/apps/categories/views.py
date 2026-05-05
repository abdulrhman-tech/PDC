"""
Categories views — hierarchical self-referencing model.
Read for all authenticated users, write for super_admin only.
"""
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
import io
import logging
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from apps.categories.models import Category, SubCategory, CategoryAttributeSchema
from apps.products.models import Product
from apps.categories.serializers import (
    CategorySerializer, CategoryWriteSerializer, CategoryTreeSerializer,
    CategoryFlatSerializer, CategoryAttributeSchemaSerializer,
    SubCategorySerializer, SubCategoryWriteSerializer,
)
from apps.integrations.translate_views import translate_text_core, TranslateError

logger = logging.getLogger(__name__)


# Regex character classes used to detect actual language content.
# Arabic Unicode block U+0600-U+06FF covers the letters we care about;
# we rely on PostgreSQL's POSIX regex via Django's __regex lookup.
_ARABIC_RX = r'[\u0600-\u06FF]'
_LATIN_RX  = r'[A-Za-z]'


def _untranslated_qs():
    """
    Categories where one of the two language fields is effectively missing
    OR the two fields are swapped (Arabic text living in ``name_en`` while
    ``name_ar`` only holds an SAP code like ``DC0000000``).

    "Effectively missing" means either truly empty OR filled with content
    that doesn't contain a single letter of the expected script — e.g.
    SAP codes like ``AG8200100`` stuffed into ``name_ar`` are NOT a real
    Arabic translation, so the row still needs translating.

    Three repairable shapes are matched:
      1. needs_ar — name_ar is missing/code-only AND name_en has Latin letters.
      2. needs_en — name_en is missing/non-Latin AND name_ar has Arabic letters.
      3. swapped  — name_ar has no Arabic letters AND name_en has Arabic letters.
                    The bulk-translate action swaps the fields then fills the
                    English side from the now-correct Arabic name.
    """
    needs_ar = (
        (Q(name_ar='') | ~Q(name_ar__regex=_ARABIC_RX))
        & Q(name_en__regex=_LATIN_RX)
    )
    needs_en = (
        (Q(name_en='') | ~Q(name_en__regex=_LATIN_RX))
        & Q(name_ar__regex=_ARABIC_RX)
    )
    swapped = (
        ~Q(name_ar__regex=_ARABIC_RX)
        & Q(name_en__regex=_ARABIC_RX)
    )
    return Category.objects.filter(needs_ar | needs_en | swapped)


# ── Attribute-schema translation helpers ─────────────────────────────────
_CODE_RX = re.compile(r'^[A-Za-z0-9_]+$')


def _looks_like_code(text: str) -> bool:
    """True if the string is empty or looks like a SAP-style identifier
    (UPPER_SNAKE_CASE / no spaces / digits, e.g. ``RECTIFIED_OR_NOT_RECTIFIED``).
    Such values are not real human labels and should be replaced."""
    t = (text or '').strip()
    if not t:
        return True
    if ' ' in t:
        return False
    return bool(_CODE_RX.match(t))


# Short uppercase tokens that should stay as acronyms even in title-cased
# output (countries, codes, units commonly written ALL-CAPS).
_KNOWN_ACRONYMS = {
    'KSA', 'USA', 'UAE', 'UK', 'EU', 'UN', 'AC', 'DC', 'TV',
    'PVC', 'MDF', 'HDF', 'CNC', 'LED', 'LCD', 'PCS', 'KG', 'MM',
    'CM', 'M2', 'M3', 'QTY', 'SKU', 'SAP', 'ID', 'API', 'URL',
    'VAT', 'GST', 'GDP',
}


def _humanize_code(text: str) -> str:
    """Turn ``RECTIFIED_OR_NOT_RECTIFIED`` / ``commission_color`` / ``KSA_COST``
    into ``Rectified Or Not Rectified`` / ``Commission Color`` / ``KSA Cost``.
    Returns ``text`` unchanged if it already looks human-readable."""
    t = (text or '').strip()
    if not t:
        return ''
    if not _looks_like_code(t):
        return t

    # When the WHOLE token is upper-snake (typical for SAP codes), every
    # part is uppercase so we can't use "is upper" as a signal. Title-case
    # each word, preserving only well-known acronyms.
    all_upper = t.replace('_', '').replace('-', '').isupper()
    parts = re.split(r'[_\-\s]+', t)
    out = []
    for p in parts:
        if not p:
            continue
        if all_upper:
            out.append(p if p in _KNOWN_ACRONYMS else p.capitalize())
        else:
            # Mixed-case input: preserve any token already in upper as acronym
            if p.isupper():
                out.append(p)
            else:
                out.append(p.capitalize())
    return ' '.join(out)


def _untranslated_attrs_qs():
    """Attribute schemas whose Arabic label is missing OR whose English label
    is empty / still a raw code (UPPER_SNAKE_CASE)."""
    needs_ar = Q(field_label_ar='') | ~Q(field_label_ar__regex=_ARABIC_RX)
    needs_en = (
        Q(field_label_en='')
        | Q(field_label_en__regex=r'^[A-Z][A-Z0-9_]*$')  # all-upper code
    )
    return CategoryAttributeSchema.objects.filter(needs_ar | needs_en).distinct()


class IsSuperAdminOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.is_authenticated and request.user.role == 'super_admin'


_CAT_FLAT_KEY = 'categories:flat:v2'
_CAT_TREE_KEY = 'categories:tree:v2'
_CAT_CACHE_TTL = 300   # 5 minutes — invalidated immediately on any write


def _invalidate_category_cache():
    from django.core.cache import cache
    cache.delete_many([_CAT_FLAT_KEY, _CAT_TREE_KEY])


class CategoryViewSet(viewsets.ModelViewSet):
    """
    Category CRUD + tree/flat/attributes/children actions.
    Endpoints:
      GET  /categories/tree/             — full recursive tree (root only nodes with nested children)
      GET  /categories/flat/             — flat list with path_ar/path_en (for dropdowns)
      GET  /categories/{id}/attributes/  — attribute schemas for this category
      POST /categories/{id}/attributes/  — add attribute schema
      GET  /categories/{id}/children/    — direct children
      POST /categories/{id}/children/    — add child category
    """
    permission_classes = [IsSuperAdminOrReadOnly]

    def perform_create(self, serializer):
        super().perform_create(serializer)
        _invalidate_category_cache()

    def perform_update(self, serializer):
        super().perform_update(serializer)
        _invalidate_category_cache()

    def perform_destroy(self, instance):
        super().perform_destroy(instance)
        _invalidate_category_cache()

    def get_queryset(self):
        qs = Category.objects.prefetch_related(
            'children', 'subcategories', 'attribute_schemas'
        ).select_related('parent').order_by('sort_order', 'order', 'name_ar')

        user = self.request.user
        if user.is_authenticated and getattr(user, 'role', None) == 'مدير_قسم':
            managed = user.get_managed_category_ids()
            qs = qs.filter(pk__in=managed) if managed else qs.none()
        return qs

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return CategoryWriteSerializer
        return CategorySerializer

    # ── Tree view ────────────────────────────────────────────────
    @action(detail=False, methods=['get'], url_path='tree')
    def tree(self, request):
        """
        Returns full hierarchical tree (root nodes with nested children).
        Optimized: uses 2 queries total (categories + attribute counts),
        builds the tree in Python — avoids the N+1 storm of the recursive
        serializer (~150 queries for a 50-node tree).
        Result is cached in Redis for 5 minutes.
        """
        from django.core.cache import cache
        from django.db.models import Count

        cached = cache.get(_CAT_TREE_KEY)
        if cached is not None:
            return Response(cached)

        # Single query for ALL categories
        cats = list(
            Category.objects.all()
            .order_by('sort_order', 'order', 'name_ar')
            .values(
                'id', 'code', 'name_ar', 'name_en', 'level', 'sort_order',
                'is_active', 'icon', 'description_ar', 'parent_id',
            )
        )

        # Single query: schemas live on root categories only
        attr_counts = {
            row['category_id']: row['c']
            for row in CategoryAttributeSchema.objects
                .values('category_id')
                .annotate(c=Count('id'))
        }

        # Index by id and group children by parent
        by_id = {c['id']: c for c in cats}
        children_map = {}
        for c in cats:
            children_map.setdefault(c['parent_id'], []).append(c)

        # Resolve each node's root ancestor (with memoization)
        root_cache = {}
        def root_id(cid):
            if cid in root_cache:
                return root_cache[cid]
            cur = by_id.get(cid)
            while cur and cur['parent_id'] is not None and cur['parent_id'] in by_id:
                cur = by_id[cur['parent_id']]
            rid = cur['id'] if cur else cid
            root_cache[cid] = rid
            return rid

        # Pass 1: compute attribute_count via root resolution (needs parent_id intact on all nodes)
        for c in cats:
            c['attribute_count'] = attr_counts.get(root_id(c['id']), 0)

        # Pass 2: attach children + counts. Keep parent_id so the edit dialog
        # can show the current parent and allow moving categories.
        for c in cats:
            kids = children_map.get(c['id'], [])
            c['children'] = kids
            c['children_count'] = len(kids)

        result = children_map.get(None, [])
        cache.set(_CAT_TREE_KEY, result, _CAT_CACHE_TTL)
        return Response(result)

    # ── Flat list (for dropdowns) ─────────────────────────────────
    @action(detail=False, methods=['get'], url_path='flat')
    def flat(self, request):
        """Returns flat list of all categories with breadcrumb paths.

        Optimized: avoids the per-row ``children.exists()`` and ``get_ancestors()``
        queries used by ``CategoryFlatSerializer`` — with ~1.4k categories
        those would explode into thousands of queries and time out. Instead
        we load every category once, build paths from an in-memory map, and
        compute ``has_children`` from a single distinct-parent query.
        Result is cached in Redis for 5 minutes to avoid the 518 KB payload
        being regenerated on every page load.
        """
        from django.core.cache import cache

        cached = cache.get(_CAT_FLAT_KEY)
        if cached is not None:
            return Response(cached)

        cats = list(Category.objects.all().values(
            'id', 'code', 'name_ar', 'name_en', 'level',
            'parent_id', 'sort_order', 'is_active',
        ))
        by_id = {c['id']: c for c in cats}
        parent_ids = set(
            Category.objects.exclude(parent_id__isnull=True)
            .values_list('parent_id', flat=True).distinct()
        )

        # Categories that have products (directly or via descendants).
        # Step 1: distinct category_ids actually referenced by Product rows.
        direct_with_products = set(
            Product.objects.exclude(category_id__isnull=True)
            .values_list('category_id', flat=True).distinct()
        )
        # Step 2: walk up parent chain so every ancestor of a leaf-with-products
        # is also marked. This way a level-1 root counts as "has products" as long
        # as any descendant under it has at least one product.
        has_products: set[int] = set()
        for cid in direct_with_products:
            node_id = cid
            seen: set[int] = set()
            while node_id is not None and node_id not in seen:
                seen.add(node_id)
                has_products.add(node_id)
                parent = by_id.get(node_id)
                node_id = parent.get('parent_id') if parent else None

        def _path(cat, field):
            parts: list[str] = []
            node = cat
            seen: set[int] = set()
            while node is not None and node['id'] not in seen:
                seen.add(node['id'])
                parts.append(node.get(field) or node.get('name_ar') or '')
                pid = node.get('parent_id')
                node = by_id.get(pid) if pid else None
            return ' > '.join(reversed(parts))

        out = []
        for c in cats:
            out.append({
                'id': c['id'],
                'code': c['code'],
                'name_ar': c['name_ar'],
                'name_en': c['name_en'],
                'level': c['level'],
                'parent': c['parent_id'],
                'sort_order': c['sort_order'],
                'is_active': c['is_active'],
                'path_ar': _path(c, 'name_ar'),
                'path_en': _path(c, 'name_en'),
                'has_children': c['id'] in parent_ids,
                'has_products': c['id'] in has_products,
            })
        # Match the previous ordering: level, sort_order, name_ar
        out.sort(key=lambda r: (r['level'], r['sort_order'] or 0, r['name_ar'] or ''))
        cache.set(_CAT_FLAT_KEY, out, _CAT_CACHE_TTL)
        return Response(out)

    # ── Attributes ────────────────────────────────────────────────
    @action(detail=True, methods=['get', 'post'])
    def attributes(self, request, pk=None):
        """
        GET: return root-ancestor's attribute schemas (inherited by all children).
        POST: add new schema to the root ancestor.
        Attributes always belong to the L1 root category; children inherit them.
        """
        category = self.get_object()

        # Walk up to find the root ancestor (level=1)
        root = category
        while root.parent_id is not None:
            root = root.parent

        if request.method == 'POST':
            if not request.user.is_authenticated or request.user.role != 'super_admin':
                return Response({'detail': 'غير مسموح.'}, status=status.HTTP_403_FORBIDDEN)
            serializer = CategoryAttributeSchemaSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save(category=root)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        schemas = CategoryAttributeSchema.objects.filter(category=root).order_by('order')
        return Response({
            'root_id': root.id,
            'root_name_ar': root.name_ar,
            'is_inherited': root.id != category.id,
            'schemas': CategoryAttributeSchemaSerializer(schemas, many=True).data,
        })

    # ── Children ──────────────────────────────────────────────────
    @action(detail=True, methods=['get', 'post'], url_path='children')
    def children(self, request, pk=None):
        """GET: list direct children. POST: add child category."""
        parent = self.get_object()
        if request.method == 'POST':
            if not request.user.is_authenticated or request.user.role != 'super_admin':
                return Response({'detail': 'غير مسموح.'}, status=status.HTTP_403_FORBIDDEN)
            if parent.level >= 5:
                return Response(
                    {'detail': 'الحد الأقصى لعمق التصنيف هو 5 مستويات.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            data = request.data.copy()
            data['parent'] = parent.pk
            serializer = CategoryWriteSerializer(data=data)
            serializer.is_valid(raise_exception=True)
            child = serializer.save()
            return Response(CategoryFlatSerializer(child).data, status=status.HTTP_201_CREATED)

        children = parent.children.all().order_by('sort_order', 'name_ar')
        return Response(CategoryFlatSerializer(children, many=True).data)

    # ── Excel Template Download ───────────────────────────────────
    @action(detail=False, methods=['get'], url_path='import-excel/template', permission_classes=[permissions.IsAuthenticated])
    def import_excel_template(self, request):
        """Download an Excel template for bulk category import."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'التصنيفات'

        gold = 'FFC8A84B'
        dark = 'FF1A1A2E'
        light_bg = 'FFFFF8E7'

        # Header row
        headers = ['code', 'name_ar', 'name_en', 'parent_code', 'sort_order']
        labels  = ['كود التصنيف *', 'الاسم بالعربية *', 'الاسم بالإنجليزية', 'كود التصنيف الأب', 'ترتيب العرض']
        req     = [True, True, False, False, False]

        header_font = Font(bold=True, color=dark, size=11, name='Arial')
        header_fill = PatternFill(fill_type='solid', fgColor=gold)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Row 1 — Arabic labels
        for col, (label, required) in enumerate(zip(labels, req), 1):
            cell = ws.cell(row=1, column=col, value=f'{label}{"" if not required else ""}')
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Row 2 — English keys (field names)
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=key)
            cell.font = Font(bold=True, color='FF555555', size=10, name='Courier New')
            cell.fill = PatternFill(fill_type='solid', fgColor='FFF5F5F5')
            cell.alignment = Alignment(horizontal='center')

        # Row 3 — Instructions
        instructions = [
            'فريد، حروف وأرقام وشرطة فقط (CERAMICS، CAT-001)',
            'الاسم الكامل بالعربية',
            'الاسم بالإنجليزية (اختياري)',
            'اتركه فارغاً للتصنيفات الرئيسية',
            '0 = الأول، 1 = الثاني...',
        ]
        for col, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=3, column=col, value=instruction)
            cell.font = Font(italic=True, color='FF888888', size=9)
            cell.fill = PatternFill(fill_type='solid', fgColor=light_bg)
            cell.alignment = Alignment(wrap_text=True, horizontal='right')

        # Sample rows
        samples = [
            ('CERAMICS',         'السيراميك والبورسلان', 'Ceramics & Porcelain', '',         1),
            ('CERAMICS-FLOOR',   'سيراميك الأرضيات',    'Floor Ceramics',       'CERAMICS',  1),
            ('CERAMICS-WALL',    'سيراميك الجدران',     'Wall Ceramics',        'CERAMICS',  2),
            ('CERAMICS-FL-MATT', 'أرضيات مات',          'Matte Floors',         'CERAMICS-FLOOR', 1),
            ('MARBLE',           'الرخام',              'Marble',               '',          2),
            ('MARBLE-ITALIAN',   'رخام إيطالي',         'Italian Marble',       'MARBLE',    1),
        ]
        sample_fill = PatternFill(fill_type='solid', fgColor='FFFAFAFA')
        alt_fill    = PatternFill(fill_type='solid', fgColor='FFFFFFFF')
        for r, row in enumerate(samples, 4):
            fill = sample_fill if r % 2 == 0 else alt_fill
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = fill
                cell.alignment = Alignment(horizontal='right' if col in (2, 3) else 'left')
                if col == 1 or col == 4:
                    cell.font = Font(name='Courier New', size=10)

        # Column widths
        for col, width in zip(['A', 'B', 'C', 'D', 'E'], [22, 28, 28, 22, 12]):
            ws.column_dimensions[col].width = width

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 40

        # Freeze header rows
        ws.freeze_panes = 'A4'

        # Instructions sheet
        ws2 = wb.create_sheet('تعليمات')
        ws2.sheet_view.rightToLeft = True
        rules = [
            ('📌 قواعد الاستيراد', None, True),
            ('', None, False),
            ('1. حقل code مطلوب وفريد — لا يُقبل تكرار الكود في الملف أو النظام.', None, False),
            ('2. parent_code يجب أن يكون موجوداً في النظام أو في الملف نفسه.', None, False),
            ('3. الحد الأقصى لعمق الشجرة هو 5 مستويات.', None, False),
            ('4. الكود يدعم: حروف إنجليزية، أرقام، وشرطة (-) فقط.', None, False),
            ('5. sort_order رقم اختياري — يحدد ترتيب العرض (0 = الأول).', None, False),
            ('6. التصنيفات الموجودة مسبقاً (بنفس الكود) تُحدَّث تلقائياً.', None, False),
            ('', None, False),
            ('💡 نصيحة: ضع التصنيفات الأبوية قبل الفرعية في الملف.', None, False),
            ('   (النظام يُرتّب تلقائياً بناءً على parent_code قبل الإنشاء)', None, False),
        ]
        for r, (text, _, bold) in enumerate(rules, 1):
            cell = ws2.cell(row=r, column=1, value=text)
            cell.font = Font(bold=bold, size=11 if bold else 10)
            if bold:
                cell.fill = PatternFill(fill_type='solid', fgColor=gold)
                cell.font = Font(bold=True, color=dark, size=13)
        ws2.column_dimensions['A'].width = 70

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="categories_import_template.xlsx"'
        return response

    # ── Excel Import ──────────────────────────────────────────────
    @action(
        detail=False, methods=['post'], url_path='import-excel',
        permission_classes=[permissions.IsAuthenticated],
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_excel(self, request):
        """
        Bulk import/update categories from an Excel file.
        Columns: code, name_ar, name_en (opt), parent_code (opt), sort_order (opt)
        Validates: no duplicate codes, parent_code exists, max 5 levels.
        Updates existing categories; creates new ones.
        """
        if not request.user.is_authenticated or request.user.role != 'super_admin':
            return Response({'detail': 'هذه العملية للمديرين العامين فقط.'}, status=status.HTTP_403_FORBIDDEN)

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'يرجى رفع ملف Excel.'}, status=status.HTTP_400_BAD_REQUEST)

        # ── Parse workbook ────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            return Response({'detail': 'ملف Excel غير صالح أو تالف.'}, status=status.HTTP_400_BAD_REQUEST)

        # Find header row (look for 'code' in first 5 rows)
        header_row_idx = None
        headers_map = {}
        for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
            for col_idx, val in enumerate(row):
                if str(val or '').strip().lower() == 'code':
                    header_row_idx = row_idx
                    break
            if header_row_idx:
                for col_idx, val in enumerate(row):
                    key = str(val or '').strip().lower()
                    if key:
                        headers_map[key] = col_idx
                break

        if header_row_idx is None:
            return Response({'detail': 'لم يُعثر على صف الترويسة. تأكد أن العمود الأول اسمه "code".'}, status=status.HTTP_400_BAD_REQUEST)

        required_cols = {'code', 'name_ar'}
        missing = required_cols - set(headers_map.keys())
        if missing:
            return Response({'detail': f'الأعمدة المطلوبة غير موجودة: {", ".join(missing)}'}, status=status.HTTP_400_BAD_REQUEST)

        # ── Read rows ─────────────────────────────────────────────
        rows_data = []
        errors = []
        seen_codes = {}

        for row_num, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), header_row_idx + 1):
            def cell(key):
                idx = headers_map.get(key)
                return str(row[idx] or '').strip() if idx is not None and row[idx] is not None else ''

            code        = cell('code').upper().replace(' ', '-')
            name_ar     = cell('name_ar')
            name_en     = cell('name_en')
            parent_code = cell('parent_code').upper().replace(' ', '-')
            try:
                sort_order = int(float(cell('sort_order'))) if cell('sort_order') else 0
            except ValueError:
                sort_order = 0

            # Skip fully empty rows
            if not code and not name_ar:
                continue

            # Validate code
            if not code:
                errors.append({'row': row_num, 'error': 'حقل code فارغ'})
                continue
            if not name_ar:
                errors.append({'row': row_num, 'code': code, 'error': 'name_ar فارغ'})
                continue

            # Check duplicate in file
            if code in seen_codes:
                errors.append({'row': row_num, 'code': code, 'error': f'الكود "{code}" مكرر في السطر {seen_codes[code]}'})
                continue

            seen_codes[code] = row_num
            rows_data.append({
                'code': code, 'name_ar': name_ar, 'name_en': name_en,
                'parent_code': parent_code, 'sort_order': sort_order,
                'row': row_num,
            })

        if not rows_data and errors:
            return Response({'detail': 'الملف يحتوي على أخطاء فقط.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        # ── Topological sort (parents before children) ────────────
        def topo_sort(rows):
            code_to_row = {r['code']: r for r in rows}
            result = []
            visited = set()

            def visit(code):
                if code in visited:
                    return
                visited.add(code)
                row = code_to_row.get(code)
                if row and row['parent_code'] and row['parent_code'] in code_to_row:
                    visit(row['parent_code'])
                if row:
                    result.append(row)

            for row in rows:
                visit(row['code'])
            return result

        ordered = topo_sort(rows_data)

        # ── Process rows in transaction ───────────────────────────
        created, updated, skipped = 0, 0, 0
        row_errors = list(errors)

        with transaction.atomic():
            # Build in-transaction lookup of all existing codes
            existing_codes = {c.code: c for c in Category.objects.all()}

            for row in ordered:
                code        = row['code']
                name_ar     = row['name_ar']
                name_en     = row['name_en']
                parent_code = row['parent_code']
                sort_order  = row['sort_order']
                row_num     = row['row']

                # Resolve parent
                parent = None
                if parent_code:
                    parent = existing_codes.get(parent_code)
                    if parent is None:
                        row_errors.append({'row': row_num, 'code': code, 'error': f'parent_code "{parent_code}" غير موجود في النظام أو في الملف'})
                        skipped += 1
                        continue

                # Compute level
                level = (parent.level + 1) if parent else 1
                if level > 5:
                    row_errors.append({'row': row_num, 'code': code, 'error': 'تجاوز الحد الأقصى للعمق (5 مستويات)'})
                    skipped += 1
                    continue

                if code in existing_codes:
                    # Update
                    cat = existing_codes[code]
                    cat.name_ar = name_ar
                    if name_en:
                        cat.name_en = name_en
                    if parent:
                        cat.parent = parent
                    cat.level = level
                    cat.sort_order = sort_order
                    cat.save(update_fields=['name_ar', 'name_en', 'parent', 'level', 'sort_order', 'updated_at'])
                    existing_codes[code] = cat
                    updated += 1
                else:
                    # Create
                    cat = Category.objects.create(
                        code=code, name_ar=name_ar, name_en=name_en,
                        parent=parent, level=level, sort_order=sort_order,
                        is_active=True,
                    )
                    existing_codes[code] = cat
                    created += 1

        return Response({
            'detail': f'اكتمل الاستيراد: {created} جديد، {updated} محدَّث، {skipped} متخطَّى',
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'errors': row_errors,
        }, status=status.HTTP_200_OK)

    # ── Legacy: subcategories (backward compat) ───────────────────
    @action(detail=True, methods=['get', 'post'], url_path='subcategories')
    def subcategories(self, request, pk=None):
        """Legacy endpoint — returns direct children as subcategory format."""
        category = self.get_object()
        if request.method == 'POST':
            if not request.user.is_authenticated or request.user.role != 'super_admin':
                return Response({'detail': 'غير مسموح.'}, status=status.HTTP_403_FORBIDDEN)
            serializer = SubCategoryWriteSerializer(
                data=request.data, context={'category': category}
            )
            serializer.is_valid(raise_exception=True)
            sub = serializer.save(category=category)
            return Response(SubCategorySerializer(sub).data, status=status.HTTP_201_CREATED)
        subs = SubCategory.objects.filter(category=category).order_by('name_ar')
        return Response(SubCategorySerializer(subs, many=True).data)

    # ── Bulk translate untranslated categories ────────────────────
    @action(
        detail=False, methods=['get'], url_path='untranslated-count',
        permission_classes=[permissions.IsAuthenticated],
    )
    def untranslated_count(self, request):
        """Returns how many categories are missing one of (name_ar, name_en)."""
        return Response({'count': _untranslated_qs().count()})

    @action(
        detail=False, methods=['post'], url_path='bulk-translate',
        permission_classes=[permissions.IsAuthenticated],
    )
    def bulk_translate(self, request):
        """
        Translate up to `limit` categories that are missing one of name_ar/name_en.
        Sequential processing to respect external translator rate limits.

        Inputs:
          limit:       int, default 20, hard-capped at 50.
          exclude_ids: list[int] (optional) — IDs the client knows already
                       failed in this session. Skipped for selection AND for
                       the returned `remaining` count, so a few persistently
                       failing items can never starve the rest of the queue.

        Writes go through `Category.objects.filter(pk=...).update(...)` so the
        hierarchy logic in `Category.save()` (level recomputation, descendant
        cascade) is intentionally NOT triggered — translation is a name-only
        write and must not touch tree shape.

        Response: {processed, succeeded, failed, remaining, errors, succeeded_ids}
        """
        if not request.user.is_authenticated or request.user.role != 'super_admin':
            return Response({'detail': 'غير مسموح.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            limit = int(request.data.get('limit', 20))
        except (TypeError, ValueError):
            limit = 20
        limit = max(1, min(limit, 50))

        raw_excludes = request.data.get('exclude_ids') or []
        if not isinstance(raw_excludes, list):
            raw_excludes = []
        exclude_ids = [int(x) for x in raw_excludes if isinstance(x, (int, str)) and str(x).isdigit()]

        qs = _untranslated_qs().exclude(id__in=exclude_ids).order_by('id')[:limit]
        succeeded = 0
        succeeded_ids: list[int] = []
        failed = 0
        errors: list[dict] = []
        now = timezone.now()

        # Same heuristic as `_untranslated_qs()`, applied per-row so a
        # category whose name_ar contains only an SAP code (no Arabic
        # letters) is treated as needing translation, not skipped.
        ar_re = re.compile(_ARABIC_RX)
        la_re = re.compile(_LATIN_RX)

        for cat in qs:
            try:
                ar = cat.name_ar or ''
                en = cat.name_en or ''
                ar_ok = bool(ar_re.search(ar))
                en_ok = bool(la_re.search(en))
                en_has_arabic = bool(ar_re.search(en))

                # Swapped fields: name_en holds Arabic text and name_ar has
                # no Arabic (typically an SAP code). Move the Arabic word
                # into name_ar first, then fall through to translate the
                # missing English side from it.
                if not ar_ok and en_has_arabic:
                    ar = en.strip()
                    Category.objects.filter(pk=cat.pk).update(
                        name_ar=ar, name_en='', updated_at=now,
                    )
                    en = ''
                    ar_ok = True
                    en_ok = False

                if not ar_ok and en_ok:
                    translated, _ = translate_text_core(en, 'en', 'ar')
                    translated = (translated or '').strip()
                    # Reject "translations" that don't actually contain Arabic
                    # letters — the LLM does this for pure-code inputs like
                    # "ASTM SCH.40" or "WJ SANDS" (returns the input back
                    # verbatim). If we accept it, _untranslated_qs() rightly
                    # keeps flagging the row, and the client loops forever.
                    if not ar_re.search(translated):
                        raise TranslateError(
                            f'No Arabic letters in result for "{en}"',
                            'لا يوجد محتوى قابل للترجمة (رمز/اختصار فقط)',
                        )
                    Category.objects.filter(pk=cat.pk).update(
                        name_ar=translated, updated_at=now,
                    )
                elif not en_ok and ar_ok:
                    translated, _ = translate_text_core(ar, 'ar', 'en')
                    translated = (translated or '').strip()
                    if not la_re.search(translated):
                        raise TranslateError(
                            f'No Latin letters in result for "{ar}"',
                            'لا يوجد محتوى قابل للترجمة (رمز/اختصار فقط)',
                        )
                    Category.objects.filter(pk=cat.pk).update(
                        name_en=translated, updated_at=now,
                    )
                else:
                    continue  # both sides valid (race with another writer)
                succeeded += 1
                succeeded_ids.append(cat.id)
            except TranslateError as exc:
                failed += 1
                if len(errors) < 10:
                    errors.append({
                        'id': cat.id, 'code': cat.code,
                        'name': cat.name_ar or cat.name_en,
                        'error': exc.friendly,
                    })
            except Exception as exc:  # pragma: no cover — defensive
                failed += 1
                logger.exception('Unexpected bulk-translate error for cat %s', cat.id)
                if len(errors) < 10:
                    errors.append({
                        'id': cat.id, 'code': cat.code,
                        'name': cat.name_ar or cat.name_en,
                        'error': str(exc)[:200],
                    })

        # `remaining` excludes the IDs the client has marked as known-failed,
        # so the loop's stop condition is honest about real progress.
        remaining = _untranslated_qs().exclude(id__in=exclude_ids).count()
        return Response({
            'processed': succeeded + failed,
            'succeeded': succeeded,
            'succeeded_ids': succeeded_ids,
            'failed': failed,
            'remaining': remaining,
            'errors': errors,
        })

    # ── Bulk-translate attribute schema labels ─────────────────────────
    @action(
        detail=False, methods=['get', 'post'], url_path='attributes-untranslated-count',
        permission_classes=[permissions.IsAuthenticated],
    )
    def attributes_untranslated_count(self, request):
        """Return how many attribute schemas still need Arabic/English labels."""
        return Response({'count': _untranslated_attrs_qs().count()})

    @action(
        detail=False, methods=['post'], url_path='bulk-translate-attributes',
        permission_classes=[permissions.IsAuthenticated],
    )
    def bulk_translate_attributes(self, request):
        """
        Translate up to `limit` attribute-schema labels (field_label_ar /
        field_label_en) that look like raw SAP codes (UPPER_SNAKE_CASE) or
        are missing the expected script.

        Inputs:
          limit:       int, default 15, hard-capped at 40.
          exclude_ids: list[int] (optional) — schema IDs the client knows
                       already failed in this session.

        Response: {processed, succeeded, succeeded_ids, failed, remaining, errors}
        """
        if not request.user.is_authenticated or request.user.role != 'super_admin':
            return Response({'detail': 'غير مسموح.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            limit = int(request.data.get('limit', 15))
        except (TypeError, ValueError):
            limit = 15
        limit = max(1, min(limit, 40))

        raw_excludes = request.data.get('exclude_ids') or []
        if not isinstance(raw_excludes, list):
            raw_excludes = []
        exclude_ids = [int(x) for x in raw_excludes if isinstance(x, (int, str)) and str(x).isdigit()]

        qs = _untranslated_attrs_qs().exclude(id__in=exclude_ids).order_by('id')[:limit]
        succeeded = 0
        succeeded_ids: list[int] = []
        failed = 0
        skipped = 0
        skipped_ids: list[int] = []
        errors: list[dict] = []

        ar_re = re.compile(_ARABIC_RX)
        la_re = re.compile(_LATIN_RX)

        for sch in qs:
            try:
                ar = (sch.field_label_ar or '').strip()
                en = (sch.field_label_en or '').strip()
                key = (sch.field_key or '').strip()

                ar_has_arabic = bool(ar_re.search(ar))
                en_is_real    = bool(en) and not _looks_like_code(en)
                en_is_code    = not en_is_real

                # A schema is "code-only" when neither label carries any real
                # natural-language content — only SAP-style identifiers like
                # ``DECORTAKM1``, ``COMMERCIALCLASS`` or ``VIT_ITEMS``. There
                # is no meaningful translation for these, so we skip them
                # silently rather than logging a scary error each time.
                if not ar_has_arabic and not en_is_real:
                    skipped += 1
                    skipped_ids.append(sch.id)
                    continue

                # Build a usable English source string. Prefer existing EN if
                # it's already human-readable; otherwise humanize the EN code,
                # the AR code, or fall back to the field key. NEVER fall back
                # to Arabic text — `field_label_en` must end up in Latin.
                if en_is_real:
                    en_source = en
                elif en:
                    en_source = _humanize_code(en)
                elif ar and not ar_has_arabic:
                    en_source = _humanize_code(ar)
                elif key:
                    en_source = _humanize_code(key)
                else:
                    en_source = ''

                # If we still have no usable English source AND the only thing
                # we have is Arabic text, translate AR→EN to populate it.
                if not en_source and ar_has_arabic:
                    translated, _ = translate_text_core(ar, 'ar', 'en')
                    translated = (translated or '').strip()
                    if not la_re.search(translated):
                        # Translation produced no Latin output → nothing
                        # meaningful to save. Treat as a skip, not a failure.
                        skipped += 1
                        skipped_ids.append(sch.id)
                        continue
                    en_source = translated

                if not en_source:
                    skipped += 1
                    skipped_ids.append(sch.id)
                    continue

                updates: dict = {}
                # Normalize field_label_en when it's empty or still a raw code
                if en_is_code:
                    updates['field_label_en'] = en_source

                # Translate to Arabic when field_label_ar lacks Arabic letters
                if not ar_has_arabic:
                    translated, _ = translate_text_core(en_source, 'en', 'ar')
                    translated = (translated or '').strip()
                    if not ar_re.search(translated):
                        # No Arabic letters returned — likely the source is a
                        # proper noun / code that the model could not localize.
                        # Save any EN normalization we did, then skip.
                        if updates:
                            CategoryAttributeSchema.objects.filter(pk=sch.pk).update(**updates)
                        skipped += 1
                        skipped_ids.append(sch.id)
                        continue
                    updates['field_label_ar'] = translated

                if not updates:
                    continue  # nothing to do (race)
                CategoryAttributeSchema.objects.filter(pk=sch.pk).update(**updates)
                succeeded += 1
                succeeded_ids.append(sch.id)
            except TranslateError as exc:
                failed += 1
                if len(errors) < 10:
                    errors.append({
                        'id': sch.id, 'key': sch.field_key,
                        'name': sch.field_label_ar or sch.field_label_en,
                        'error': exc.friendly,
                    })
            except Exception as exc:  # pragma: no cover — defensive
                failed += 1
                logger.exception('Unexpected bulk-translate-attrs error for schema %s', sch.id)
                if len(errors) < 10:
                    errors.append({
                        'id': sch.id, 'key': sch.field_key,
                        'name': sch.field_label_ar or sch.field_label_en,
                        'error': str(exc)[:200],
                    })

        # Permanently exclude code-only / unsalvageable schemas from the
        # remaining count for this session so the loop terminates cleanly.
        session_excludes = set(exclude_ids) | set(skipped_ids)
        remaining = _untranslated_attrs_qs().exclude(id__in=session_excludes).count()
        return Response({
            'processed': succeeded + failed + skipped,
            'succeeded': succeeded,
            'succeeded_ids': succeeded_ids,
            'failed': failed,
            'skipped': skipped,
            'skipped_ids': skipped_ids,
            'remaining': remaining,
            'errors': errors,
        })


class CategoryAttributeSchemaViewSet(viewsets.ModelViewSet):
    """CRUD for individual attribute schema records — super_admin only."""
    serializer_class = CategoryAttributeSchemaSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = CategoryAttributeSchema.objects.select_related('category').order_by(
        'category__sort_order', 'category__order', 'order'
    )

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticatedOrReadOnly()]
        return [permissions.IsAuthenticated()]

    def _check_admin(self, request):
        if request.user.role != 'super_admin':
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('هذه العملية للمديرين العامين فقط.')

    def create(self, request, *args, **kwargs):
        self._check_admin(request)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        self._check_admin(request)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        self._check_admin(request)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        self._check_admin(request)
        return super().destroy(request, *args, **kwargs)


class SubCategoryViewSet(viewsets.ModelViewSet):
    """Legacy: Update/delete individual subcategory records."""
    queryset = SubCategory.objects.select_related('category').order_by('name_ar')
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return SubCategoryWriteSerializer
        return SubCategorySerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        try:
            ctx['category'] = self.get_object().category
        except Exception:
            pass
        return ctx

    def _check_admin(self, request):
        if not request.user.is_authenticated or request.user.role != 'super_admin':
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('هذه العملية للمديرين العامين فقط.')

    def update(self, request, *args, **kwargs):
        self._check_admin(request)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        self._check_admin(request)
        instance = self.get_object()
        serializer = SubCategoryWriteSerializer(
            instance, data=request.data, partial=True,
            context={'category': instance.category}
        )
        serializer.is_valid(raise_exception=True)
        sub = serializer.save()
        return Response(SubCategorySerializer(sub).data)

    def destroy(self, request, *args, **kwargs):
        self._check_admin(request)
        return super().destroy(request, *args, **kwargs)
