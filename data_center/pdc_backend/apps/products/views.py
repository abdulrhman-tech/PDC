"""
Product views (DRF ViewSets) for Bayt Alebaa PDC.
Enforces role-based access and category-scoping for dept managers.
"""
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter

from apps.products.models import Product, Brand, ProductStatus, ProductSubmission, SubmissionImage, SubmissionStatus
from apps.products.serializers import (
    ProductListSerializer,
    ProductDetailSerializer,
    ProductCreateUpdateSerializer,
    BrandSerializer,
)
from apps.products.permissions import ProductPermissions
from apps.products.filters import ProductFilter


# ─── خريطة أسماء الحقول البسيطة بالعربية ───
FIELD_LABELS_AR = {
    'product_name_ar':  'الاسم (عربي)',
    'product_name_en':  'الاسم (إنجليزي)',
    'description_ar':   'الوصف (عربي)',
    'price_sar':        'السعر (ريال)',
    'sku':              'رمز المنتج',
    'status':           'الحالة',
    'brand':            'العلامة التجارية',
    'category':         'القسم',
    'subcategory':      'الفئة الفرعية',
    'origin_country':   'بلد المنشأ',
    'color':            'اللون',
    'inventory_type':   'نوع المخزون',
    'ecommerce_url':    'رابط التجارة الإلكترونية',
    'stock_quantity':   'الكمية',
    'unit':             'الوحدة',
    'warranty_months':  'مدة الضمان (شهر)',
    'is_featured':      'منتج مميز',
}


def _val_to_str(val) -> str:
    """حوّل قيمة حقل إلى نص مقروء مختصر."""
    if val is None or val == '':
        return '—'
    if isinstance(val, bool):
        return 'نعم' if val else 'لا'
    if hasattr(val, 'pk'):        # FK object
        return str(val)
    s = str(val)
    # اختصر النصوص الطويلة (مثل الأوصاف)
    return s[:60] + '…' if len(s) > 60 else s


def _build_product_diff(instance, validated_data: dict) -> str:
    """
    قارن الحقول المُعدَّلة بين القيمة القديمة والجديدة.
    يغطي الحقول البسيطة وأيضًا كل مفاتيح attributes.
    """
    lines = []

    # ── حقول بسيطة ──
    for field, label in FIELD_LABELS_AR.items():
        if field not in validated_data:
            continue
        old_val = getattr(instance, field, None)
        new_val = validated_data[field]
        old_str = _val_to_str(old_val)
        new_str = _val_to_str(new_val)
        if old_str != new_str:
            lines.append(f'{label}: [{old_str}] ← [{new_str}]')

    # ── حقل attributes (JSON) ──
    if 'attributes' in validated_data:
        old_attrs = getattr(instance, 'attributes', {}) or {}
        new_attrs = validated_data['attributes'] or {}
        all_keys = set(old_attrs) | set(new_attrs)
        for key in sorted(all_keys):
            old_a = _val_to_str(old_attrs.get(key))
            new_a = _val_to_str(new_attrs.get(key))
            if old_a != new_a:
                lines.append(f'{key}: [{old_a}] ← [{new_a}]')

    if not lines:
        return 'لم يتم رصد تغييرات على الحقول المتابَعة'
    return '\n'.join(lines)


class ProductViewSet(viewsets.ModelViewSet):
    """
    Full CRUD for products.
    - GET /api/v1/products/             → list
    - POST /api/v1/products/            → create (dept manager + super admin only)
    - GET /api/v1/products/{id}/        → detail
    - PUT/PATCH /api/v1/products/{id}/  → update
    - DELETE /api/v1/products/{id}/     → delete (super admin only)
    - POST /api/v1/products/{id}/publish/ → publish (super admin only)
    - POST /api/v1/products/{id}/generate-description/ → AI description
    - POST /api/v1/products/{id}/generate-image/ → AI decorative image (Celery)
    """
    permission_classes = [ProductPermissions]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = ProductFilter
    search_fields = ['product_name_ar', 'product_name_en', 'sku', 'description_ar']
    ordering_fields = ['created_at', 'updated_at', 'product_name_ar', 'price_sar', 'status']
    ordering = ['-created_at']

    def get_queryset(self):
        user = self.request.user
        qs = Product.objects.select_related(
            'category', 'category__parent', 'category__parent__parent',
            'subcategory', 'brand', 'created_by', 'updated_by'
        ).prefetch_related('images')

        # مدير قسم: only sees products inside any of his assigned categories
        # (and their descendant subtrees).
        if user.is_authenticated and user.is_dept_manager:
            managed = user.get_managed_category_ids()
            qs = qs.filter(category_id__in=managed) if managed else qs.none()

        return qs

    def get_serializer_class(self):
        if self.action == 'list':
            return ProductListSerializer
        if self.action in ['create', 'update', 'partial_update']:
            return ProductCreateUpdateSerializer
        return ProductDetailSerializer

    def _create_approval_request(self, product, user, request_type, previous_status='', request=None, diff=''):
        """Helper: create an approval request after a dept manager action."""
        from apps.approvals.models import ProductApprovalRequest
        from apps.logs.utils import log_action
        # إلغاء أي طلب معلّق سابق لنفس المنتج
        ProductApprovalRequest.objects.filter(
            product=product, status='pending'
        ).update(status='rejected', reviewer_notes='إلغاء تلقائي — طلب جديد')

        approval = ProductApprovalRequest.objects.create(
            product=product,
            submitted_by=user,
            request_type=request_type,
            previous_status=previous_status,
            status='pending',
        )
        # سجّل على المنتج نفسه مع تفاصيل التغييرات
        action_key = 'create_product' if request_type == 'new_product' else 'update_product'
        if request_type == 'new_product':
            detail_msg = f'إضافة منتج جديد {product.sku} — قيد انتظار الموافقة'
        else:
            detail_msg = f'تعديل منتج {product.sku} — قيد انتظار الموافقة'
            if diff:
                detail_msg += f'\n{diff}'
        log_action(user, action_key, product, detail_msg, request)
        # سجّل طلب الاعتماد
        log_action(user, 'approval_submit', approval,
                   f'طلب {"إضافة" if request_type == "new_product" else "تعديل"} منتج {product.sku}', request)

    def perform_create(self, serializer):
        from apps.logs.utils import log_action
        user = self.request.user
        req = self.request
        if user.is_dept_manager:
            product = serializer.save(status='قيد_المراجعة', created_by=user, updated_by=user)
            self._create_approval_request(product, user, 'new_product', previous_status='', request=req)
        else:
            product = serializer.save(created_by=user, updated_by=user)
            log_action(user, 'create_product', product, f'إضافة منتج {product.sku}', req)

    def perform_update(self, serializer):
        from apps.logs.utils import log_action
        user = self.request.user
        req = self.request
        instance = serializer.instance
        previous_status = instance.status

        # احسب الـ diff قبل الحفظ (بينما القيم القديمة لا تزال في الـ instance)
        diff = _build_product_diff(instance, serializer.validated_data)

        if user.is_dept_manager:
            product = serializer.save(status='قيد_المراجعة', updated_by=user)
            self._create_approval_request(
                product, user, 'edit_product',
                previous_status=previous_status, request=req, diff=diff
            )
        else:
            product = serializer.save(updated_by=user)
            detail_msg = f'تعديل منتج {product.sku}\n{diff}'
            log_action(user, 'update_product', product, detail_msg, req)

    @action(detail=True, methods=['post'], url_path='publish')
    def publish(self, request, pk=None):
        """Publish a product — Super Admin only."""
        if not request.user.can_publish_product():
            return Response(
                {'detail': 'ليس لديك صلاحية نشر المنتجات.'},
                status=status.HTTP_403_FORBIDDEN
            )
        product = self.get_object()
        if not product.images.filter(image_type='main', status='approved').exists():
            return Response(
                {'detail': 'يجب رفع صورة رئيسية معتمدة قبل النشر.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        product.status = ProductStatus.ACTIVE
        product.published_at = timezone.now()
        product.updated_by = request.user
        product.save(update_fields=['status', 'published_at', 'updated_by', 'updated_at'])

        # Log the action
        from apps.logs.utils import log_action
        log_action(request.user, 'publish_product', product, f'نشر المنتج {product.sku}')

        return Response({'detail': 'تم نشر المنتج بنجاح.', 'status': product.status})

    @action(detail=True, methods=['post'], url_path='generate-description')
    def generate_description(self, request, pk=None):
        """Generate AI bilingual description (AR + EN) via OpenAI."""
        product = self.get_object()
        try:
            from apps.integrations.openai_service import generate_product_description
            descriptions = generate_product_description(product)
            if descriptions.get('description_ar'):
                product.description_ar = descriptions['description_ar']
            if descriptions.get('description_en'):
                product.description_en = descriptions['description_en']
            product.save(update_fields=['description_ar', 'description_en'])
            return Response({
                'description_ar': product.description_ar,
                'description_en': product.description_en,
            })
        except Exception as e:
            err_str = str(e)
            if 'quota' in err_str.lower() or 'rate' in err_str.lower() or '429' in err_str:
                user_msg = 'تجاوزنا الحد المسموح به مؤقتاً. يرجى المحاولة بعد دقيقة.'
            elif 'auth' in err_str.lower() or '401' in err_str:
                user_msg = 'مفتاح الذكاء الاصطناعي غير صالح. يرجى مراجعة الإعدادات.'
            else:
                user_msg = 'فشل توليد الوصف. يرجى المحاولة مرة أخرى.'
            import logging
            logging.getLogger(__name__).error(f"OpenAI description error for {product.sku}: {e}")
            return Response(
                {'detail': user_msg},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )

    @action(detail=True, methods=['post'], url_path='generate-image')
    def generate_image(self, request, pk=None):
        """Queue AI decorative image generation via Celery."""
        product = self.get_object()
        from apps.integrations.tasks import generate_decorative_image_task
        task = generate_decorative_image_task.delay(product.id)
        return Response({
            'detail': 'تم إضافة طلب توليد الصورة إلى قائمة المعالجة.',
            'task_id': task.id
        })

    @action(detail=True, methods=['get'], url_path='attributes')
    def attributes(self, request, pk=None):
        """Return product attributes with their schema definition."""
        product = self.get_object()
        from apps.categories.models import CategoryAttributeSchema
        schema = CategoryAttributeSchema.objects.filter(
            category=product.category
        ).order_by('order')
        data = {
            'attributes': product.attributes,
            'schema': [
                {
                    'key': s.field_key,
                    'label_ar': s.field_label_ar,
                    'type': s.field_type,
                    'options': s.options,
                    'required': s.is_required,
                    'unit': s.unit,
                }
                for s in schema
            ]
        }
        return Response(data)

    @action(detail=True, methods=['get'], url_path='images')
    def images(self, request, pk=None):
        """GET → list all images for this product."""
        product = self.get_object()
        from apps.images.models import ProductImage
        from apps.products.serializers import ProductImageSerializer
        qs = ProductImage.objects.filter(product=product).order_by('image_type', 'order')
        return Response(ProductImageSerializer(qs, many=True).data)

    @action(detail=True, methods=['post'], url_path='images/upload')
    def upload_image(self, request, pk=None):
        """
        POST multipart/form-data with 'file' + 'image_type'.
        Uploads file to R2 via server-side, saves ProductImage record.
        """
        product = self.get_object()
        file = request.FILES.get('file')
        image_type = request.data.get('image_type', 'gallery')

        if not file:
            return Response({'detail': 'لم يُرسل أي ملف'}, status=status.HTTP_400_BAD_REQUEST)

        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/webp', 'image/gif']
        if file.content_type not in allowed_types:
            return Response({'detail': 'نوع الملف غير مدعوم'}, status=status.HTTP_400_BAD_REQUEST)

        if file.size > 10 * 1024 * 1024:
            return Response({'detail': 'حجم الملف أكبر من 10 ميجابايت'}, status=status.HTTP_400_BAD_REQUEST)

        import uuid
        ext = file.name.rsplit('.', 1)[-1].lower() if '.' in file.name else 'jpg'
        unique_name = f"{uuid.uuid4().hex[:10]}.{ext}"
        r2_key = f"products/{product.category.slug}/{product.sku}/{image_type}/{unique_name}"

        try:
            from apps.integrations.r2_client import upload_bytes
            file_bytes = file.read()
            r2_url = upload_bytes(r2_key, file_bytes, content_type=file.content_type)
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"R2 upload error: {e}")
            return Response({'detail': 'فشل رفع الملف إلى التخزين'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        from apps.images.models import ProductImage
        from apps.products.serializers import ProductImageSerializer
        img = ProductImage.objects.create(
            product=product,
            image_type=image_type,
            r2_key=r2_key,
            r2_url=r2_url,
            original_filename=file.name,
            file_size_kb=round(file.size / 1024),
            status='approved',
            uploaded_by=request.user if request.user.is_authenticated else None,
        )
        return Response(ProductImageSerializer(img).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['delete', 'patch'], url_path=r'images/(?P<image_id>\d+)')
    def image_detail(self, request, pk=None, image_id=None):
        """
        DELETE → remove an image record (+ R2 object)
        PATCH  → update image_type or order
        """
        from apps.images.models import ProductImage
        from apps.products.serializers import ProductImageSerializer

        product = self.get_object()
        try:
            img = ProductImage.objects.get(pk=image_id, product=product)
        except ProductImage.DoesNotExist:
            return Response({'detail': 'الصورة غير موجودة'}, status=status.HTTP_404_NOT_FOUND)

        if request.method == 'DELETE':
            # Try to delete from R2 as well
            try:
                from apps.integrations.r2_client import delete_object
                delete_object(img.r2_key)
            except Exception:
                pass
            img.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

        # PATCH
        for field in ['image_type', 'order', 'status']:
            if field in request.data:
                setattr(img, field, request.data[field])
        img.save()
        return Response(ProductImageSerializer(img).data)

    @action(detail=False, methods=['get'], url_path='import-excel/template', permission_classes=[permissions.IsAuthenticated])
    def import_excel_template(self, request):
        """Return a category-specific downloadable Excel template."""
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from apps.categories.models import Category as Cat, CategoryAttributeSchema

        category_id = request.query_params.get('category_id')
        cat_obj = None
        attr_schemas = []
        if category_id:
            try:
                cat_obj = Cat.objects.get(pk=category_id)
                attr_schemas = list(CategoryAttributeSchema.objects.filter(category=cat_obj).order_by('is_required', 'order').reverse())
                # is_required first: reverse so required come first
                attr_schemas = sorted(attr_schemas, key=lambda a: (0 if a.is_required else 1, a.order))
            except Cat.DoesNotExist:
                pass

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'منتجات - {cat_obj.name_ar}' if cat_obj else 'منتجات'

        # ── Basic columns ──
        base_keys = ['sku', 'product_name_ar', 'product_name_en',
                     'brand_name', 'origin_country', 'color', 'inventory_type',
                     'status', 'price_sar', 'ecommerce_url', 'description_ar']
        base_labels = ['رمز المنتج *', 'الاسم (عربي) *', 'الاسم (إنجليزي)',
                       'الماركة', 'بلد المنشأ', 'اللون', 'نوع المخزون (دوري/ستوك)',
                       'الحالة (نشط/مسودة)', 'السعر (ريال)', 'رابط المتجر', 'الوصف (عربي)']

        # ── Dynamic attribute columns ──
        attr_keys = [f'attr__{a.field_key}' for a in attr_schemas]
        attr_labels = []
        for a in attr_schemas:
            label = a.field_label_ar
            if a.is_required:
                label += ' *'
            if a.unit:
                label += f' ({a.unit})'
            if a.options:
                label += f' [{"/".join(a.options[:4])}]'
            attr_labels.append(label)

        # If category selected, drop 'category_slug' from required; add it as fixed
        if cat_obj:
            all_keys = base_keys + attr_keys
            all_labels = base_labels + attr_labels
        else:
            all_keys = ['category_slug'] + base_keys + attr_keys
            all_labels = ['التصنيف (slug) *'] + base_labels + attr_labels

        # ── Styles ──
        base_fill = PatternFill(start_color='1A2636', end_color='1A2636', fill_type='solid')
        attr_fill = PatternFill(start_color='0D3320', end_color='0D3320', fill_type='solid')
        base_font = Font(bold=True, color='C8A84B', size=10)
        attr_font = Font(bold=True, color='4ADE80', size=10)
        req_font  = Font(bold=True, color='FF6B6B', size=10)
        thin = Side(style='thin', color='334455')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        base_col_count = len(all_keys) - len(attr_keys)

        for col, (key, label) in enumerate(zip(all_keys, all_labels), 1):
            is_attr = col > base_col_count
            is_req = '*' in label

            # Row 1: Arabic label
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = attr_fill if is_attr else base_fill
            cell.font = attr_font if is_attr else (req_font if is_req else base_font)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

            # Row 2: machine key
            ws.cell(row=2, column=col, value=key)

        ws.row_dimensions[1].height = 40

        # ── Sample row ──
        sample_base = ['SKU-001', 'اسم المنتج', 'Product Name',
                       'اسم الماركة', 'السعودية', 'بيج', 'دوري',
                       'نشط', '500', '', 'وصف المنتج']
        if not cat_obj:
            sample_base = [cat_obj.slug if cat_obj else 'ceramics-tiles'] + sample_base
        sample_attrs = []
        for a in attr_schemas:
            if a.options:
                sample_attrs.append(a.options[0])
            elif a.field_type == 'boolean':
                sample_attrs.append('نعم')
            elif a.field_type == 'number':
                sample_attrs.append('0')
            else:
                sample_attrs.append('')
        sample = sample_base + sample_attrs
        for col, val in enumerate(sample, 1):
            cell = ws.cell(row=3, column=col, value=val)
            cell.alignment = Alignment(horizontal='right')

        # ── Column widths ──
        for col in range(1, len(all_keys) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22

        # ── Notes ──
        note_text = '* الحقول الحمراء مطلوبة | الأعمدة الخضراء = سمات ديناميكية | ابدأ بيانات من الصف 3 | لا تحذف الصف 2'
        ws.cell(row=4, column=1, value=note_text).font = Font(color='888888', italic=True, size=8)

        if cat_obj:
            ws.cell(row=5, column=1, value=f'القسم: {cat_obj.name_ar} | slug: {cat_obj.slug}').font = Font(color='C8A84B', bold=True, size=9)

        # ── Options reference sheet ──
        has_opts = [a for a in attr_schemas if a.options]
        if has_opts:
            ws_opts = wb.create_sheet(title='الخيارات المتاحة')
            ws_opts.cell(row=1, column=1, value='الحقل').font = Font(bold=True, color='C8A84B')
            ws_opts.cell(row=1, column=2, value='الخيارات المتاحة').font = Font(bold=True, color='C8A84B')
            ws_opts.column_dimensions['A'].width = 25
            ws_opts.column_dimensions['B'].width = 50
            for i, a in enumerate(has_opts, 2):
                ws_opts.cell(row=i, column=1, value=a.field_label_ar)
                ws_opts.cell(row=i, column=2, value=' | '.join(a.options))

        fname = f'template_{cat_obj.slug}.xlsx' if cat_obj else 'products_import_template.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='import-excel', permission_classes=[permissions.IsAuthenticated])
    def import_excel(self, request):
        """Bulk import products from Excel.

        Supports two formats automatically:
        1) PDC template — row 1 = Arabic labels, row 2 = machine keys, data from row 3.
        2) SAP export — row 1 = SAP headers (Material_No, Material_Description,
           Material_No.<Field>...), data from row 2. Category is resolved by
           matching `Material_No.Material Group No` to `Category.code`.
        """
        from apps.logs.utils import log_action
        import openpyxl
        from apps.categories.models import Category, CategoryAttributeSchema

        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({'detail': 'لم يُرفع أي ملف'}, status=status.HTTP_400_BAD_REQUEST)

        category_id = request.data.get('category_id')
        fixed_cat = None
        attr_schemas = {}
        if category_id:
            try:
                fixed_cat = Category.objects.get(pk=category_id)
                for a in CategoryAttributeSchema.objects.filter(category=fixed_cat):
                    attr_schemas[f'attr__{a.field_key}'] = a.field_key
            except Category.DoesNotExist:
                return Response({'detail': 'القسم المحدد غير موجود'}, status=status.HTTP_400_BAD_REQUEST)

        # NOTE: We deliberately use read_only=False even though it loads the
        # whole workbook into memory. Reason: openpyxl's read_only mode trusts
        # the workbook's `<dimension>` XML metadata, which is frequently stale
        # in files exported by SAP / BI tools / scripts (e.g. declares "A1:Z250"
        # while the sheet actually contains 5000 rows). With stale metadata,
        # iter_rows() in read-only mode silently stops at the declared bound
        # and rows are lost without any error. Non-read-only mode always uses
        # the real used range. Since we immediately materialize everything via
        # list(...) below, read_only mode wasn't giving us a memory benefit
        # anyway — a typical 5000-row product sheet is well under 50 MB.
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=False, data_only=True)
            ws = wb.active
        except Exception:
            return Response({'detail': 'ملف Excel غير صالح'}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response({'detail': 'الملف فارغ أو لا يحتوي على بيانات'}, status=status.HTTP_400_BAD_REQUEST)

        # ── Format detection ──
        # SAP exports have headers like Material_No, Material_Description, and
        # Material_No.Material Group No on row 1. Normalize each header (strip
        # non-alphanumeric, lowercase) so variants like "Material Description"
        # or "Material No" still detect correctly.
        import re as _re
        first_row = [str(c).strip() if c is not None else '' for c in rows[0]]
        first_row_norm = [_re.sub(r'[^a-z0-9]', '', c.lower()) for c in first_row]
        is_sap_format = (
            'materialno' in first_row_norm
            and any(n.startswith('materialdescription') for n in first_row_norm)
        )

        if is_sap_format:
            return self._import_sap_excel(request, rows, first_row)

        if len(rows) < 3:
            return Response({'detail': 'الملف فارغ أو لا يحتوي على بيانات'}, status=status.HTTP_400_BAD_REQUEST)

        # Row 2 is the machine-readable key row
        keys = [str(k).strip() if k else '' for k in rows[1]]

        # Cache lookups
        cats = {c.slug: c for c in Category.objects.all()}
        brands_map = {}
        for b in Brand.objects.all():
            if b.name_ar:
                brands_map[b.name_ar] = b
            brands_map[b.name] = b

        created = []
        errors = []

        for row_num, row in enumerate(rows[2:], start=3):
            row_data = {}
            for i, k in enumerate(keys):
                row_data[k] = str(row[i]).strip() if i < len(row) and row[i] is not None else ''

            sku = row_data.get('sku', '').strip()
            name_ar = row_data.get('product_name_ar', '').strip()

            # Skip completely empty rows
            if not sku and not name_ar:
                continue

            row_errors = []
            if not sku:
                row_errors.append('رمز المنتج (SKU) مطلوب')
            if not name_ar:
                row_errors.append('اسم المنتج (عربي) مطلوب')

            # Resolve category
            if fixed_cat:
                cat_obj = fixed_cat
            else:
                cat_slug = row_data.get('category_slug', '').strip()
                if not cat_slug:
                    row_errors.append('التصنيف مطلوب')
                    cat_obj = None
                elif cat_slug not in cats:
                    row_errors.append(f'التصنيف "{cat_slug}" غير موجود')
                    cat_obj = None
                else:
                    cat_obj = cats[cat_slug]

            if sku and Product.objects.filter(sku=sku).exists():
                row_errors.append(f'الـ SKU "{sku}" موجود مسبقاً')

            if row_errors:
                errors.append({'row': row_num, 'sku': sku or '—', 'errors': row_errors})
                continue

            try:
                price_raw = row_data.get('price_sar', '')
                try:
                    price = float(price_raw) if price_raw else None
                except ValueError:
                    price = None

                create_kwargs = dict(
                    sku=sku,
                    product_name_ar=name_ar,
                    product_name_en=row_data.get('product_name_en', ''),
                    category=cat_obj,
                    description_ar=row_data.get('description_ar', ''),
                    status=row_data.get('status', 'مسودة') or 'مسودة',
                    inventory_type=row_data.get('inventory_type', 'دوري') or 'دوري',
                    origin_country=row_data.get('origin_country', ''),
                    color=row_data.get('color', ''),
                    ecommerce_url=row_data.get('ecommerce_url', ''),
                    created_by=request.user,
                    updated_by=request.user,
                )
                if price is not None:
                    create_kwargs['price_sar'] = price

                # Brand
                brand_name = row_data.get('brand_name', '').strip()
                if brand_name and brand_name in brands_map:
                    create_kwargs['brand'] = brands_map[brand_name]

                # Dynamic attributes
                attributes = {}
                for col_key, attr_key in attr_schemas.items():
                    val = row_data.get(col_key, '').strip()
                    if val:
                        attributes[attr_key] = val
                if attributes:
                    create_kwargs['attributes'] = attributes

                product = Product.objects.create(**create_kwargs)
                created.append({'row': row_num, 'sku': sku, 'name': name_ar, 'id': product.id})
            except Exception as e:
                errors.append({'row': row_num, 'sku': sku, 'errors': [str(e)]})

        # Diagnostic counters help operators detect "silent loss" cases
        # like the one caused by stale workbook <dimension> metadata.
        total_rows_in_file = len(rows)
        data_rows_seen = max(0, total_rows_in_file - 2)  # rows 1+2 are header
        log_action(request.user, 'excel_import', None,
                   f'استيراد Excel: {len(created)} منتج مضاف، '
                   f'{len(errors)} خطأ، إجمالي صفوف الملف: {total_rows_in_file}',
                   request)

        return Response({
            'created_count': len(created),
            'error_count': len(errors),
            'total_rows_in_file': total_rows_in_file,
            'data_rows_seen': data_rows_seen,
            'created': created,
            'errors': errors,
        })

    def _import_sap_excel(self, request, rows, headers):
        """Import products from a SAP-format Excel.

        Expected layout:
          Row 1 (headers): Material_No, Material_Description,
                           Material_No.Material Group No, Material_No.<X>, ...
          Rows 2+        : data
        Category is matched on `Category.code == Material Group No`.
        Other `Material_No.<X>` columns are mapped to attributes by fuzzy
        matching against the per-category `CategoryAttributeSchema.field_key`
        (case- and separator-insensitive).
        """
        import re
        from apps.logs.utils import log_action
        from apps.categories.models import Category, CategoryAttributeSchema

        def _norm(s: str) -> str:
            """Normalize identifier: lowercase, strip non-alphanumeric."""
            return re.sub(r'[^a-z0-9]', '', (s or '').lower())

        # Map header index → role / normalized name. Only columns whose
        # header starts with `Material_No.` are eligible to become product
        # attributes; bare `Material_No` and `Material_Description` are
        # special, and any other unrelated columns are ignored.
        sku_idx = None
        name_idx = None
        group_no_idx = None
        brand_idx = None
        origin_idx = None
        color_idx = None
        attr_cols = []  # list of (col_idx, normalized_name, original_label)

        for i, h in enumerate(headers):
            if not h:
                continue
            h_lower = h.lower().strip()
            h_norm = _norm(h)
            if h_norm == 'materialno':
                sku_idx = i
                continue
            if h_norm == 'materialdescription':
                name_idx = i
                continue

            # Only `Material_No.<X>` (or `Material No <X>`) columns count
            # as SAP fields. Anything else is ignored to avoid accidental
            # matches with unrelated workbook columns.
            label = None
            if h_lower.startswith('material_no.'):
                label = h[len('material_no.'):].strip()
            elif h_lower.startswith('material no.'):
                label = h[len('material no.'):].strip()
            elif h_lower.startswith('material_no '):
                label = h[len('material_no '):].strip()
            if label is None:
                continue

            label_norm = _norm(label)

            if label_norm == 'materialgroupno':
                group_no_idx = i
                continue
            if label_norm == 'brand':
                brand_idx = i
                continue
            if label_norm in ('brandorigin', 'cityoforigin'):
                if origin_idx is None:
                    origin_idx = i
                continue
            if label_norm == 'itemcolor':
                color_idx = i
                continue

            attr_cols.append((i, label_norm, label))

        if sku_idx is None:
            return Response(
                {'detail': 'لم يُعثر على عمود Material_No'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if group_no_idx is None:
            return Response(
                {'detail': 'لم يُعثر على عمود Material_No.Material Group No'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Preload categories indexed by code (case-insensitive) and id
        cats_by_code = {}
        cats_by_id = {}
        for c in Category.objects.exclude(code=''):
            cats_by_code[c.code.strip().upper()] = c
            cats_by_id[c.id] = c
        for c in Category.objects.filter(code=''):
            cats_by_id[c.id] = c

        # Preload attribute schemas grouped by category id (direct schemas only)
        direct_schemas_by_cat = {}
        for s in CategoryAttributeSchema.objects.all():
            direct_schemas_by_cat.setdefault(s.category_id, {})[_norm(s.field_key)] = s

        # Cache for "effective" schemas including those inherited from
        # ancestor categories. Walks the parent chain and merges schemas;
        # closer (deeper) categories override the same normalized key.
        effective_schemas_cache: dict[int, dict] = {}

        def _effective_schemas(cat) -> dict:
            if cat is None:
                return {}
            if cat.id in effective_schemas_cache:
                return effective_schemas_cache[cat.id]
            chain = []
            node = cat
            seen = set()
            while node is not None and node.id not in seen:
                seen.add(node.id)
                chain.append(node)
                parent_id = getattr(node, 'parent_id', None)
                node = cats_by_id.get(parent_id) if parent_id else None
            merged: dict = {}
            # Walk root-first so deeper category schemas override shallower
            for n in reversed(chain):
                merged.update(direct_schemas_by_cat.get(n.id, {}))
            effective_schemas_cache[cat.id] = merged
            return merged

        # Brand cache
        brands_map = {}
        for b in Brand.objects.all():
            if b.name_ar:
                brands_map[b.name_ar.strip().lower()] = b
            brands_map[b.name.strip().lower()] = b

        def _cell(row, idx):
            if idx is None or idx >= len(row):
                return ''
            v = row[idx]
            if v is None:
                return ''
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            return str(v).strip()

        created = []
        updated = []
        errors = []

        for row_num, row in enumerate(rows[1:], start=2):
            sku = _cell(row, sku_idx)
            name_ar = _cell(row, name_idx)
            group_no = _cell(row, group_no_idx)

            if not sku and not name_ar and not group_no:
                continue  # blank row

            row_errors = []
            if not sku:
                row_errors.append('Material_No فارغ')
            if not name_ar:
                # Fall back: use SKU as name if description is missing
                name_ar = sku

            cat_obj = None
            if not group_no:
                row_errors.append('Material Group No فارغ')
            else:
                cat_obj = cats_by_code.get(group_no.strip().upper())
                if cat_obj is None:
                    row_errors.append(
                        f'لا يوجد قسم بكود "{group_no}" — قم بمزامنة الأقسام من SAP أولاً'
                    )

            if row_errors:
                errors.append({'row': row_num, 'sku': sku or '—',
                               'errors': row_errors})
                continue

            # Build attributes by fuzzy-matching column labels to schema field_keys
            attributes = {}
            cat_schema = _effective_schemas(cat_obj)
            for col_idx, col_norm, col_label in attr_cols:
                if col_idx == group_no_idx:
                    continue
                val = _cell(row, col_idx)
                if not val:
                    continue
                schema = cat_schema.get(col_norm)
                if schema is None:
                    continue  # column not part of this category's schema
                attributes[schema.field_key] = val

            try:
                existing = Product.objects.filter(sku=sku).first()
                base_kwargs = dict(
                    product_name_ar=name_ar,
                    category=cat_obj,
                    updated_by=request.user,
                )
                # Optional simple fields if present in the file
                brand_name = _cell(row, brand_idx).lower() if brand_idx is not None else ''
                if brand_name and brand_name in brands_map:
                    base_kwargs['brand'] = brands_map[brand_name]

                origin = _cell(row, origin_idx)
                if origin:
                    base_kwargs['origin_country'] = origin

                color = _cell(row, color_idx)
                if color:
                    base_kwargs['color'] = color

                if existing:
                    for k, v in base_kwargs.items():
                        setattr(existing, k, v)
                    # Merge attributes (new values overwrite, others kept)
                    merged = dict(existing.attributes or {})
                    merged.update(attributes)
                    existing.attributes = merged
                    existing.save()
                    updated.append({'row': row_num, 'sku': sku,
                                    'name': name_ar, 'id': existing.id})
                else:
                    base_kwargs.update(
                        sku=sku,
                        attributes=attributes,
                        status='مسودة',
                        inventory_type='دوري',
                        created_by=request.user,
                    )
                    product = Product.objects.create(**base_kwargs)
                    created.append({'row': row_num, 'sku': sku,
                                    'name': name_ar, 'id': product.id})
            except Exception as e:
                errors.append({'row': row_num, 'sku': sku, 'errors': [str(e)]})

        # Diagnostic counters help operators detect "silent loss" cases
        # like the one caused by stale workbook <dimension> metadata.
        total_rows_in_file = len(rows)
        data_rows_seen = max(0, total_rows_in_file - 1)  # row 1 = headers
        log_action(
            request.user, 'excel_import', None,
            f'استيراد SAP: {len(created)} جديد، {len(updated)} محدّث، '
            f'{len(errors)} خطأ، إجمالي صفوف الملف: {total_rows_in_file}',
            request,
        )

        return Response({
            'format': 'sap',
            'created_count': len(created),
            'updated_count': len(updated),
            'error_count': len(errors),
            'total_rows_in_file': total_rows_in_file,
            'data_rows_seen': data_rows_seen,
            'created': created,
            'updated': updated,
            'errors': errors,
        })

    @action(detail=True, methods=['post'], url_path='images/generate-url', permission_classes=[permissions.IsAuthenticated])
    def generate_upload_url(self, request, pk=None):
        """Generate a presigned R2 URL for direct client-side upload."""
        product = self.get_object()
        image_type = request.data.get('image_type', 'gallery')
        filename = request.data.get('filename', 'image.jpg')
        content_type = request.data.get('content_type', 'image/jpeg')

        from apps.integrations.r2_client import generate_presigned_url
        import uuid
        unique_name = f"{uuid.uuid4().hex[:8]}_{filename}"
        r2_key = (
            f"products/{product.category.slug}/{product.sku}/"
            f"{image_type}/{unique_name}"
        )
        url = generate_presigned_url(r2_key, content_type)
        return Response({'upload_url': url, 'r2_key': r2_key})


class BrandViewSet(viewsets.ModelViewSet):
    serializer_class = BrandSerializer
    search_fields = ['name', 'name_ar']
    ordering = ['name']

    def get_queryset(self):
        # Authenticated users (admin/manager) see all brands; public sees only active
        if self.request.user and self.request.user.is_authenticated:
            return Brand.objects.all().order_by('name_ar', 'name')
        return Brand.objects.filter(is_active=True).order_by('name_ar', 'name')

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]


class ProductSubmissionViewSet(viewsets.ModelViewSet):
    """
    Handles visitor product submissions + manager/admin review workflow.
    POST   /products/submissions/          — public (visitor creates)
    GET    /products/submissions/          — staff only
    PATCH  /products/submissions/{id}/     — staff only
    POST   /products/submissions/{id}/submit_for_approval/  — manager
    POST   /products/submissions/{id}/approve/              — admin
    POST   /products/submissions/{id}/reject/               — admin
    """
    from apps.products.serializers import ProductSubmissionSerializer
    serializer_class = ProductSubmissionSerializer
    http_method_names = ['get', 'post', 'patch', 'delete']

    def get_serializer_class(self):
        from apps.products.serializers import ProductSubmissionSerializer
        return ProductSubmissionSerializer

    def get_queryset(self):
        user = self.request.user
        qs = ProductSubmission.objects.select_related('category', 'assigned_manager', 'product').prefetch_related('images')
        if not user.is_authenticated:
            return ProductSubmission.objects.none()
        if user.is_super_admin:
            return qs.all()
        if user.is_dept_manager:
            managed = user.get_managed_category_ids()
            return qs.filter(category_id__in=managed) if managed else qs.none()
        return qs.filter(assigned_manager=user)

    def get_permissions(self):
        if self.action == 'create':
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def create(self, request, *args, **kwargs):
        """
        Public endpoint: visitor submits a new product request.
        Accepts multipart/form-data with optional image files.
        """
        from apps.products.serializers import ProductSubmissionSerializer
        from apps.logs.utils import log_action
        import uuid

        data = request.data
        serializer = ProductSubmissionSerializer(data={
            'sku':             data.get('sku', ''),
            'category':        data.get('category'),
            'product_name_ar': data.get('product_name_ar', ''),
            'submitter_name':  data.get('submitter_name', ''),
            'submitter_email': data.get('submitter_email', ''),
        })
        serializer.is_valid(raise_exception=True)
        submission = serializer.save()

        # Auto-assign to department manager whose department matches category
        if submission.category:
            from apps.users.models import User
            manager = User.objects.filter(
                role='مدير_قسم', department=submission.category, is_active=True
            ).first()
            if manager:
                submission.assigned_manager = manager
                submission.status = SubmissionStatus.IN_REVIEW
                submission.save(update_fields=['assigned_manager', 'status'])

        # Handle uploaded images
        images = request.FILES.getlist('images')
        if images:
            try:
                from apps.integrations.r2_client import upload_bytes
                for img_file in images[:10]:  # max 10 images
                    ext = img_file.name.rsplit('.', 1)[-1] if '.' in img_file.name else 'jpg'
                    r2_key = f'submissions/{submission.id}/{uuid.uuid4().hex}.{ext}'
                    r2_url = upload_bytes(r2_key, img_file.read(), content_type=img_file.content_type)
                    SubmissionImage.objects.create(submission=submission, r2_key=r2_key, r2_url=r2_url)
            except Exception as e:
                import logging
                logging.getLogger(__name__).error(f'Submission image upload error: {e}')

        log_action(
            user=request.user if request.user.is_authenticated else None,
            action='submission_created',
            obj=submission,
            details=f'طلب جديد من {submission.submitter_name} ({submission.submitter_email}): {submission.product_name_ar}',
            request=request,
        )
        return Response(ProductSubmissionSerializer(submission).data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        """Manager updates submission details."""
        from apps.logs.utils import log_action
        instance = self.get_object()
        allowed_fields = ['sku', 'category', 'product_name_ar', 'manager_notes', 'admin_notes', 'extra_data']
        data = {k: v for k, v in request.data.items() if k in allowed_fields}
        serializer = self.get_serializer(instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        log_action(request.user, 'submission_updated', instance, 'تحديث بيانات الطلب', request)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='submit_for_approval')
    def submit_for_approval(self, request, pk=None):
        """Manager submits to admin for final approval."""
        from apps.logs.utils import log_action
        submission = self.get_object()
        if submission.status not in (SubmissionStatus.PENDING, SubmissionStatus.IN_REVIEW):
            return Response({'detail': 'الطلب ليس في حالة تسمح بالإرسال للموافقة'}, status=status.HTTP_400_BAD_REQUEST)
        submission.status = SubmissionStatus.PENDING_APPROVAL
        submission.save(update_fields=['status'])
        log_action(request.user, 'submission_submitted_for_approval', submission, 'أُرسل للموافقة النهائية', request)
        return Response({'status': 'pending_approval'})

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Admin approves → creates actual Product and transfers submission images."""
        from apps.logs.utils import log_action
        from apps.images.models import ProductImage, ImageType, ImageStatus
        submission = self.get_object()
        if submission.status != SubmissionStatus.PENDING_APPROVAL:
            return Response({'detail': 'الطلب غير جاهز للاعتماد'}, status=status.HTTP_400_BAD_REQUEST)

        # Create the product — merge extra_data fields filled by the manager
        extra = submission.extra_data or {}
        try:
            create_kwargs = dict(
                sku=submission.sku or f'SUB-{submission.id}',
                product_name_ar=submission.product_name_ar,
                product_name_en=extra.get('product_name_en', ''),
                category=submission.category,
                status=ProductStatus.ACTIVE,
                description_ar=extra.get('description_ar', ''),
                color=extra.get('color', ''),
                origin_country=extra.get('origin_country', ''),
                inventory_type=extra.get('inventory_type', 'دوري'),
                ecommerce_url=extra.get('ecommerce_url', ''),
                attributes=extra.get('attributes', {}),
                created_by=request.user,
                updated_by=request.user,
            )
            # Optional FK fields
            if extra.get('subcategory'):
                from apps.categories.models import SubCategory
                try:
                    create_kwargs['subcategory'] = SubCategory.objects.get(pk=extra['subcategory'])
                except SubCategory.DoesNotExist:
                    pass
            if extra.get('brand'):
                try:
                    create_kwargs['brand'] = Brand.objects.get(pk=extra['brand'])
                except Brand.DoesNotExist:
                    pass
            product = Product.objects.create(**create_kwargs)
            submission.product = product
        except Exception as e:
            return Response({'detail': f'خطأ في إنشاء المنتج: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        # Transfer submission images → ProductImage
        submission_images = submission.images.all()
        for order, sub_img in enumerate(submission_images):
            img_type = ImageType.MAIN if order == 0 else ImageType.GALLERY
            ProductImage.objects.create(
                product=product,
                image_type=img_type,
                r2_key=sub_img.r2_key,
                r2_url=sub_img.r2_url,
                status=ImageStatus.APPROVED,
                order=order,
                uploaded_by=request.user,
            )

        submission.status = SubmissionStatus.APPROVED
        submission.save(update_fields=['status', 'product'])
        log_action(request.user, 'submission_approved', submission, f'اعتُمد وأُنشئ المنتج {product.sku}', request)
        return Response({'status': 'approved', 'product_id': product.id})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Admin rejects the submission."""
        from apps.logs.utils import log_action
        submission = self.get_object()
        reason = request.data.get('reason', '')
        submission.status = SubmissionStatus.REJECTED
        if reason:
            submission.admin_notes = reason
        submission.save(update_fields=['status', 'admin_notes'])
        log_action(request.user, 'submission_rejected', submission, f'رُفض: {reason}', request)
        return Response({'status': 'rejected'})
